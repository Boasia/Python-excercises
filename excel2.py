from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side

#ustawiam obramowanie komórki na cienkie:
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#ustawiam obramowanie komórki na grube:
thick_border = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))

#Trzowę obiekt klasy Workbook:
book = Workbook()
sheet = book.active

#startujemy z wpisuwaniem tabliczki mnożenia od komórki (2,2):
row_no=2
column_no = 2

#pętla do wypełnienia tabliczki mnożenia:
for row in range(1, 11):

    #wypełniam całą kolumnę:
    for col in range(1, 11):
        #poza wypełnieniem komórek dodaje cienka obramówke komórki:
        sheet.cell(row_no, column_no, row * col).border = thin_border
        row_no += 1

    # po wypełnieniu kolumny przechodzę do kolejnej:
    column_no += 1
    #zeruje wartośc zmiennej, by znów zaczął uzupełniać od góry do dołu:
    row_no = 2



#formatuję wizualnie krawędzie komórki na grube i wypełniam je liczbami 1-10:
for z in range(1,12):
    sheet.cell(1, z,z-1).border = thick_border
    sheet.cell(z,1,z-1).border = thick_border

#tworzę plik xlsx:
book.save('ex1.xlsx')
